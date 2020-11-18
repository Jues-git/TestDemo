import time

import openpyxl


class meth_excel:
    def setup(self):
        self.xl = openpyxl.Workbook()
        # 获取当前表单sheet
        self.sheet = self.xl.active
        # 创建表单sheet
        self.newsheet = self.xl.create_sheet('testcase')
        # 打开表单
        self.test = openpyxl.load_workbook(r'C:\Users\Jues\PycharmProjects\pythonProject\test.xlsx')
        # 获取当前表单case
        self.case = self.test.active
        # 获取当前时间戳
        time.strftime('%Y-%m-%d %H:%M:%S')
        # '2020-11-18 14:02:16'
        # 根据表单名字获取
        self.sheet1 = self.xl.get_sheet_by_name('sheet1')
        self.sheet2 = self.xl['sheet2']
        # 获取行数
        self.sheet.max_row
        self.sheet.max_column

    def write_sheet(self):
        # 在当前表单第一行，第一列写入数据
        self.sheet.cell(row=1, column=1, value='test')
        self.sheet['A1'] = 'test'
        self.sheet.cell(row=1, column=1).value = 'test'
        # 在新建表单sheet写入数据
        self.newsheet.cell(row=1, column=1, value='test1')
        # 写入多个数据，在第一列写入多个数据
        rows = [self.sheet.cell(row=i, column=1, value='test%d' % i) for i in range(2, 100)]
        # 逐行写入
        tableValues = [['张学友', 15201062100, 18, '测试数据！'], ['李雷', 15201062598, 19, '测试数据！'],
                       ['Marry', 15201062191, 28, '测试数据！']]
        self.sheet.append(tableValues)
    def read_sheet(self):
        # 读取A4值
        read_A = self.case['A4'].value
        read_At = self.case['A1':'C4']
        print(read_A, read_At)
        # 连续读取
        self.case.iter_rows(2, 100, 1, 10)  # 最小行2 最大行100 ，最小列1 ，最大列10  返回一个生成器
        self.case.iter_cols(1, 10, 2, 5)   # 最小列1，最大列10，最小行2，最大行5 ，返回一个生成器

        # 读取指定行数
        row1 = [row for row in self.sheet.rows][1]  # self.sheet.rows 返回所有行数，生成器
        #  row1 = (<Cell 'sheet11'.A2>, <Cell 'sheet11'.B2>)
        col1 = [col for col in self.sheet.columns][1]


